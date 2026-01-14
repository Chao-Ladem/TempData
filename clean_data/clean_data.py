"""
清理附件五-危險貨物一覽表和有限數量例外.xlsx
處理 Excel 合併儲存格造成的 Unnamed 欄位問題
一頁一頁處理，每頁單獨寫入新 Excel
"""

import pandas as pd
import numpy as np
import re
from pathlib import Path
from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter


def fix_merged_cells_final(df):
    """
    Unnamed: X只能向左找前置非Unnamed欄位
    處理Excel橫向合併儲存格的標準邏輯
    """
    cols = df.columns.tolist()
    
    for i, col in enumerate(cols):
        # 檢查是否為Unnamed欄位
        if re.match(r'^Unnamed:\s*\d+$', str(col)):
            # 向左找第一個非Unnamed欄位
            target_col = None
            for j in range(i-1, -1, -1):
                if not re.match(r'^Unnamed:\s*\d+$', str(cols[j])):
                    target_col = cols[j]
                    break
            
            if target_col:
                # 將Unnamed非空值移至前置目標欄位
                mask = df[col].notna() & df[target_col].isna()
                df.loc[mask, target_col] = df.loc[mask, col]
                if mask.sum() > 0:
                    print(f"  ✓ {col} → {target_col}: 移轉 {mask.sum()} 筆資料")
            else:
                print(f"  ⚠ {col}: 無前置目標欄位")
    
    # 統一刪除所有Unnamed欄位
    unnamed_cols = df.columns[df.columns.str.contains('^Unnamed:', na=False)]
    if len(unnamed_cols) > 0:
        df.drop(columns=unnamed_cols, inplace=True)
        print(f"  🗑 刪除 {len(unnamed_cols)} 個Unnamed欄位")
    
    return df


def read_sheet_with_headers(input_path, sheet_name):
    """
    使用 openpyxl 讀取活頁，保留標題行
    返回包含標題的 DataFrame
    """
    wb = load_workbook(input_path, read_only=True, data_only=True)
    ws = wb[sheet_name]
    
    # 讀取所有數據（包含標題）
    data = []
    for row in ws.iter_rows(values_only=True):
        data.append(row)
    
    wb.close()
    
    if len(data) < 2:
        return None, None, None
    
    # 前兩行是標題
    header_row1 = data[0]
    header_row2 = data[1]
    
    # 從第三行開始是數據
    data_rows = data[2:] if len(data) > 2 else []
    
    return header_row1, header_row2, data_rows


def create_dataframe_with_headers(header_row1, header_row2, data_rows):
    """
    使用標題行和數據行建立 DataFrame
    """
    # 確定最大欄位數
    max_cols = max(
        len(header_row1) if header_row1 else 0,
        len(header_row2) if header_row2 else 0,
        max(len(row) for row in data_rows) if data_rows else 0
    )
    
    # 建立欄位名稱（合併兩行標題）
    new_columns = []
    for i in range(max_cols):
        h1 = str(header_row1[i]) if header_row1 and i < len(header_row1) and header_row1[i] is not None else ''
        h2 = str(header_row2[i]) if header_row2 and i < len(header_row2) and header_row2[i] is not None else ''
        
        h1 = h1.strip()
        h2 = h2.strip()
        
        if h1:
            # 如果第二行有內容且不同，合併它們
            if h2 and h2 != h1 and 'Unnamed' not in h2:
                col_name = f"{h1}\n{h2}"
            else:
                col_name = h1
        elif h2:
            col_name = h2
        else:
            col_name = f"Unnamed: {i}"
        
        new_columns.append(col_name)
    
    # 標準化數據行（確保每行長度一致）
    normalized_rows = []
    for row in data_rows:
        row_list = list(row) if isinstance(row, (tuple, list)) else [row]
        while len(row_list) < max_cols:
            row_list.append(None)
        normalized_rows.append(row_list)
    
    # 建立 DataFrame
    df = pd.DataFrame(normalized_rows, columns=new_columns)
    
    return df


def clean_sheet(input_path, sheet_name):
    """
    清理單個活頁的資料
    返回清理後的 DataFrame 和標題行
    """
    print(f"\n處理活頁: {sheet_name}")
    
    try:
        # 讀取標題和數據
        header_row1, header_row2, data_rows = read_sheet_with_headers(input_path, sheet_name)
        
        if header_row1 is None or header_row2 is None:
            print(f"  ⚠ 活頁 {sheet_name} 標題不足，跳過")
            return None, None, None
        
        if not data_rows:
            print(f"  ⚠ 活頁 {sheet_name} 無數據，跳過")
            return None, None, None
        
        # 建立 DataFrame
        df = create_dataframe_with_headers(header_row1, header_row2, data_rows)
        
        print(f"  📊 原始資料: {len(df)} 筆, {len(df.columns)} 欄位")
        
        # 處理 Unnamed 欄位
        df = fix_merged_cells_final(df)
        
        print(f"  ✅ 處理完成: {len(df)} 筆, {len(df.columns)} 欄位")
        
        return df, header_row1, header_row2
        
    except Exception as e:
        print(f"  ✗ 處理活頁 {sheet_name} 時發生錯誤: {e}")
        import traceback
        traceback.print_exc()
        return None, None, None


def write_sheet_to_excel(wb, sheet_name, df, header_row1, header_row2):
    """
    將清理後的資料寫入 Excel 活頁
    保留原始標題格式
    """
    # 創建活頁
    ws = wb.create_sheet(title=sheet_name)
    
    # 寫入第一行標題
    if header_row1:
        for col_idx, value in enumerate(header_row1, 1):
            ws.cell(row=1, column=col_idx, value=value)
    
    # 寫入第二行標題
    if header_row2:
        for col_idx, value in enumerate(header_row2, 1):
            ws.cell(row=2, column=col_idx, value=value)
    
    # 寫入數據（從第三行開始）
    if df is not None and len(df) > 0:
        # 確保欄位對齊
        max_cols = max(len(header_row1) if header_row1 else 0, 
                      len(header_row2) if header_row2 else 0,
                      len(df.columns))
        
        # 寫入欄位名稱（第三行，如果需要）
        # 但我們已經有標題了，所以直接寫數據
        for row_idx, (_, row_data) in enumerate(df.iterrows(), start=3):
            for col_idx in range(max_cols):
                if col_idx < len(df.columns):
                    value = row_data.iloc[col_idx]
                    ws.cell(row=row_idx, column=col_idx + 1, value=value)


def main(input_path, output_path):
    """
    主處理函數
    """
    print("=" * 60)
    print("開始處理附件五-危險貨物一覽表和有限數量例外.xlsx")
    print("=" * 60)
    
    # 使用 openpyxl 讀取活頁名稱
    wb_input = load_workbook(input_path, read_only=True, data_only=True)
    sheet_names = wb_input.sheetnames
    wb_input.close()
    
    print(f"\n發現 {len(sheet_names)} 個活頁")
    print(f"活頁列表: {', '.join(sheet_names[:10])}{'...' if len(sheet_names) > 10 else ''}")
    
    # 創建新的 Excel 文件
    wb_output = Workbook()
    # 刪除默認的 Sheet
    if 'Sheet' in wb_output.sheetnames:
        wb_output.remove(wb_output['Sheet'])
    
    # 處理每個活頁
    success_count = 0
    for idx, sheet_name in enumerate(sheet_names, 1):
        try:
            # 清理資料
            df_clean, header_row1, header_row2 = clean_sheet(input_path, sheet_name)
            
            if df_clean is not None:
                # 寫入新 Excel
                write_sheet_to_excel(wb_output, sheet_name, df_clean, header_row1, header_row2)
                success_count += 1
                print(f"  ✓ 完成 ({idx}/{len(sheet_names)})")
            else:
                print(f"  ⚠ 活頁 {sheet_name} 無有效資料，跳過")
                
        except Exception as e:
            print(f"  ✗ 處理活頁 {sheet_name} 時發生錯誤: {e}")
            continue
    
    # 儲存到新的 Excel 文件
    print(f"\n儲存到: {output_path}")
    wb_output.save(output_path)
    
    print("\n" + "=" * 60)
    print("處理完成！")
    print("=" * 60)
    print(f"輸出檔案: {output_path}")
    print(f"成功處理: {success_count}/{len(sheet_names)} 個活頁")


def copy_headers_and_format(input_path, output_path):
    """
    複製 "Table 2" 的前兩行到其他所有活頁
    並設置第一欄為 0000 格式，自動調整欄寬，置中對齊
    """
    print("=" * 60)
    print("複製 Table 2 標題並格式化所有活頁")
    print("=" * 60)
    
    print("讀取 Excel 文件...")
    wb = load_workbook(input_path, data_only=True)
    
    # 獲取 "Table 2" 的前兩行
    if "Table 2" not in wb.sheetnames:
        print("❌ 找不到 'Table 2' 活頁！")
        wb.close()
        return
    
    ws_template = wb["Table 2"]
    
    # 讀取前兩行
    header_row1 = []
    header_row2 = []
    
    for col in range(1, ws_template.max_column + 1):
        cell1 = ws_template.cell(row=1, column=col)
        cell2 = ws_template.cell(row=2, column=col)
        header_row1.append(cell1.value)
        header_row2.append(cell2.value)
    
    print(f"✓ 讀取 'Table 2' 前兩行，共 {len(header_row1)} 欄")
    print(f"  第一行: {header_row1[:5]}...")
    print(f"  第二行: {header_row2[:5]}...")
    
    # 獲取所有活頁名稱
    all_sheets = wb.sheetnames
    print(f"\n發現 {len(all_sheets)} 個活頁")
    
    # 更新所有活頁的前兩行（除了 Table 2）
    updated_count = 0
    for sheet_name in all_sheets:
        if sheet_name == "Table 2":
            print(f"  ⏭ 跳過 'Table 2'（模板活頁）")
            continue
        
        ws = wb[sheet_name]
        print(f"  📝 更新 '{sheet_name}'...")
        
        # 先取消前兩行的所有合併儲存格
        merged_ranges_to_remove = []
        for merged_range in list(ws.merged_cells.ranges):
            # 檢查是否與前兩行重疊
            min_row, min_col, max_row, max_col = merged_range.bounds
            if min_row <= 2:  # 如果合併範圍包含前兩行
                merged_ranges_to_remove.append(merged_range)
        
        # 移除合併儲存格
        for merged_range in merged_ranges_to_remove:
            ws.unmerge_cells(str(merged_range))
        
        # 清除前兩行的值
        for row in [1, 2]:
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                if cell.value is not None:
                    cell.value = None
        
        # 寫入新的前兩行
        for col_idx, (val1, val2) in enumerate(zip(header_row1, header_row2), start=1):
            cell1 = ws.cell(row=1, column=col_idx, value=val1)
            cell2 = ws.cell(row=2, column=col_idx, value=val2)
            
            # 第一欄設置為 0000 格式（4位數，不足補零）
            if col_idx == 1:
                cell1.number_format = '0000'
                cell2.number_format = '0000'
        
        # 為第一欄的所有數據行設置格式
        for row in range(3, ws.max_row + 1):
            cell = ws.cell(row=row, column=1)
            if cell.value is not None:
                # 如果是數字，設置格式
                try:
                    num_value = float(cell.value)
                    cell.number_format = '0000'
                except (ValueError, TypeError):
                    pass
        
        # 設置所有欄位自動調整欄寬和置中對齊
        # 計算每欄的最大寬度
        column_widths = {}
        for col in range(1, ws.max_column + 1):
            max_width = 0
            for row in range(1, ws.max_row + 1):
                cell = ws.cell(row=row, column=col)
                if cell.value is not None:
                    # 計算文字長度（中文字算2個字符寬度）
                    cell_value = str(cell.value)
                    # 簡單估算：中文字符算2，其他算1
                    width = sum(2 if ord(c) > 127 else 1 for c in cell_value)
                    max_width = max(max_width, width)
            
            # 設置欄寬（最小8，最大50，加上一些邊距）
            column_widths[col] = min(max(max_width + 2, 8), 50)
            # 使用 get_column_letter 將列號轉換為列字母
            col_letter = get_column_letter(col)
            ws.column_dimensions[col_letter].width = column_widths[col]
        
        # 設置所有儲存格置中對齊
        center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        for row in range(1, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                cell.alignment = center_alignment
        
        updated_count += 1
        print(f"    ✓ 完成（欄寬已調整，已置中）")
    
    print(f"\n儲存到: {output_path}")
    wb.save(output_path)
    wb.close()
    
    print(f"\n✅ 完成！已更新 {updated_count} 個活頁的前兩行標題")


def merge_all_sheets(input_path, output_path, max_columns=11):
    """
    合併所有活頁到一個活頁
    保留前兩行標題，從第三行開始合併所有數據
    只保留 A~K 欄（11欄）
    """
    print("=" * 60)
    print("合併所有活頁到一個活頁")
    print("=" * 60)
    
    print("讀取 Excel 文件...")
    wb = load_workbook(input_path, data_only=True)
    
    # 獲取所有活頁名稱
    all_sheets = wb.sheetnames
    print(f"發現 {len(all_sheets)} 個活頁")
    
    # 獲取第一個活頁的前兩行作為標題（假設所有活頁標題相同）
    if not all_sheets:
        print("❌ 沒有活頁可以處理！")
        wb.close()
        return
    
    ws_first = wb[all_sheets[0]]
    
    # 讀取前兩行標題
    header_row1 = []
    header_row2 = []
    for col in range(1, max_columns + 1):
        cell1 = ws_first.cell(row=1, column=col)
        cell2 = ws_first.cell(row=2, column=col)
        header_row1.append(cell1.value)
        header_row2.append(cell2.value)
    
    print(f"✓ 讀取標題，共 {len(header_row1)} 欄")
    
    # 收集所有活頁的數據（從第三行開始）
    all_data_rows = []
    total_rows = 0
    
    for sheet_name in all_sheets:
        ws = wb[sheet_name]
        print(f"  📖 讀取 '{sheet_name}'...")
        
        # 從第三行開始讀取數據
        sheet_data = []
        for row in range(3, ws.max_row + 1):
            row_data = []
            for col in range(1, max_columns + 1):
                cell = ws.cell(row=row, column=col)
                row_data.append(cell.value)
            # 只添加非空行（至少有一個非空值）
            if any(val is not None for val in row_data):
                sheet_data.append(row_data)
                all_data_rows.append(row_data)
        
        print(f"    ✓ 讀取 {len(sheet_data)} 筆資料")
        total_rows += len(sheet_data)
    
    print(f"\n總共收集 {total_rows} 筆資料")
    
    # 創建新的工作簿
    wb_output = Workbook()
    ws_output = wb_output.active
    ws_output.title = "合併資料"
    
    # 寫入標題行
    print("\n寫入標題行...")
    for col_idx, (val1, val2) in enumerate(zip(header_row1, header_row2), start=1):
        cell1 = ws_output.cell(row=1, column=col_idx, value=val1)
        cell2 = ws_output.cell(row=2, column=col_idx, value=val2)
        
        # 第一欄設置為 0000 格式
        if col_idx == 1:
            cell1.number_format = '0000'
            cell2.number_format = '0000'
    
    # 寫入數據（從第三行開始）
    print("寫入數據...")
    for row_idx, row_data in enumerate(all_data_rows, start=3):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws_output.cell(row=row_idx, column=col_idx, value=value)
            
            # 第一欄設置為 0000 格式（如果是數字）
            if col_idx == 1 and value is not None:
                try:
                    num_value = float(value)
                    cell.number_format = '0000'
                except (ValueError, TypeError):
                    pass
    
    # 設置欄寬和對齊
    print("設置格式...")
    
    # 計算每欄的最大寬度
    column_widths = {}
    for col in range(1, max_columns + 1):
        max_width = 0
        for row in range(1, ws_output.max_row + 1):
            cell = ws_output.cell(row=row, column=col)
            if cell.value is not None:
                cell_value = str(cell.value)
                width = sum(2 if ord(c) > 127 else 1 for c in cell_value)
                max_width = max(max_width, width)
        
        column_widths[col] = min(max(max_width + 2, 8), 50)
        col_letter = get_column_letter(col)
        ws_output.column_dimensions[col_letter].width = column_widths[col]
    
    # 設置所有儲存格置中對齊
    center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    for row in range(1, ws_output.max_row + 1):
        for col in range(1, max_columns + 1):
            cell = ws_output.cell(row=row, column=col)
            cell.alignment = center_alignment
    
    # 儲存文件
    print(f"\n儲存到: {output_path}")
    wb_output.save(output_path)
    wb.close()
    
    print("\n" + "=" * 60)
    print("合併完成！")
    print("=" * 60)
    print(f"輸出檔案: {output_path}")
    print(f"總筆數: {total_rows}")
    print(f"總欄位數: {max_columns}")


if __name__ == "__main__":


    # 注意!!
    # 下面要一段一段跑，不要一次跑完
    # main(()
    # copy_headers_and_format()
    # merge_all_sheets()
    # 要一段一段跑!!!!


    # 輸入和輸出路徑
    input_path = Path(r"D:\Code\中華精測\sds\doc\附件五-危險貨物一覽表和有限數量例外.xlsx")
    output_path = Path(r"D:\Code\中華精測\sds\doc\附件五-危險貨物一覽表和有限數量例外_clean.xlsx")
    # main(input_path, output_path)  # 清理 Unnamed 欄位

    ## ========================================================
    #   手動調整第一個活頁
    #   - 活頁名稱: Table 2
    #   調整內容:
    #       - 標題
    #       - 第一欄自訂格式0000
    ## ========================================================
    
    # 複製標題並格式化
    copy_input = Path(r"D:\Code\中華精測\sds\doc\附件五-危險貨物一覽表和有限數量例外_clean.xlsx")
    copy_output = Path(r"D:\Code\中華精測\sds\doc\附件五-危險貨物一覽表和有限數量例外_統一標題.xlsx")
    # copy_headers_and_format(copy_input, copy_output)

    # 合併所有活頁
    final_input = Path(r"D:\Code\中華精測\sds\doc\附件五-危險貨物一覽表和有限數量例外_統一標題.xlsx")
    final_output = Path(r"D:\Code\中華精測\sds\doc\附件五-危險貨物一覽表和有限數量例外_final.xlsx")
    merge_all_sheets(final_input, final_output, max_columns=11)